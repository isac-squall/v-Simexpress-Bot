# -*- coding: utf-8 -*-
import os
import time
from pathlib import Path
import shutil
import glob
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import argparse
from openpyxl import Workbook

load_dotenv(override=True)

USUARIO = os.getenv("SIMEXPRESS_USUARIO")
SENHA = os.getenv("SIMEXPRESS_SENHA")
DOWNLOAD_PATH = os.getenv("DOWNLOAD_PATH", str(Path.cwd() / "downloads"))
EXCEL_PATH = os.getenv("EXCEL_PATH", str(Path(__file__).parent / "pedidos.xlsx"))


def _pedidos_do_env():
    _pedidos_raw = os.getenv("PEDIDOS_LOTE", "123456\\n234567\\n345678")
    if "\\n" in _pedidos_raw and "\n" not in _pedidos_raw:
        pedidos = [p.strip() for p in _pedidos_raw.split('\\n') if p.strip()]
    else:
        # Suporte a vírgulas ou quebras de linha
        pedidos = []
        for line in _pedidos_raw.splitlines():
            pedidos.extend([p.strip() for p in line.split(',') if p.strip()])
    return '\n'.join(pedidos)


def _carregar_pedidos_do_arquivo(caminho_arquivo):
    try:
        if caminho_arquivo.lower().endswith('.csv'):
            df = pd.read_csv(caminho_arquivo)
        else:
            df = pd.read_excel(caminho_arquivo)
        # Colunas de fallback padrão
        colunas_candidatas = [
            "Pedido", "pedido", "Pedidos", "pedidos",
            "Order", "order", "Número do Pedido", "numero_pedido"
        ]
        coluna_escolhida = None
        for c in colunas_candidatas:
            if c in df.columns:
                coluna_escolhida = c
                break

        if coluna_escolhida is None:
            if len(df.columns) == 0:
                raise ValueError("A planilha Excel não possui colunas")
            coluna_escolhida = df.columns[0]

        pedidos = df[coluna_escolhida].dropna().astype(str).str.strip()
        pedidos = [p for p in pedidos if p]
        if not pedidos:
            raise ValueError(f"Nenhum pedido válido encontrado na coluna '{coluna_escolhida}'")

        return '\n'.join(pedidos)
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {caminho_arquivo}. Usando PEDIDOS_LOTE do .env como fallback")
        return _pedidos_do_env()
    except Exception as e:
        print(f"Erro ao ler arquivo {caminho_arquivo}: {e}. Usando PEDIDOS_LOTE do .env como fallback")
        return _pedidos_do_env()


PEDIDOS_LOTE = _pedidos_do_env()

XPATH_USUARIO = os.getenv("SIMEXPRESS_XPATH_USUARIO", "")
XPATH_SENHA = os.getenv("SIMEXPRESS_XPATH_SENHA", "")
XPATH_ENTRAR = os.getenv("SIMEXPRESS_XPATH_ENTRAR", "")

if not USUARIO or not SENHA:
    raise RuntimeError("Defina SIMEXPRESS_USUARIO e SIMEXPRESS_SENHA no .env")

URL = "https://simexpress.com.br/"


def gerar_relatorio_excel(logs, download_path, pedidos_lista):
    wb = Workbook()
    
    # Aba Logs
    ws_logs = wb.active
    ws_logs.title = "Logs de Execução"
    ws_logs.append(["Timestamp", "Mensagem"])
    for log_entry in logs:
        # Separar timestamp e mensagem
        if " " in log_entry and log_entry.count(" ") >= 2:
            timestamp = log_entry.split(" ", 2)[0] + " " + log_entry.split(" ", 2)[1]
            mensagem = log_entry.split(" ", 2)[2]
        else:
            timestamp = ""
            mensagem = log_entry
        ws_logs.append([timestamp, mensagem])
    
    # Aba Dados do CSV (se existir)
    csv_path = Path(download_path) / "SIMEXPRESS_LOTE.csv"
    if csv_path.exists():
        try:
            df = pd.read_csv(csv_path)
            ws_dados = wb.create_sheet("Dados CSV")
            # Escrever cabeçalhos
            for col_num, col_name in enumerate(df.columns, 1):
                ws_dados.cell(row=1, column=col_num, value=col_name)
            # Escrever dados
            for row_num, row_data in enumerate(df.itertuples(index=False), 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_dados.cell(row=row_num, column=col_num, value=value)
        except Exception as e:
            print(f"Erro ao ler CSV para Excel: {e}")
    
    # Aba Screenshot (se existir)
    screenshot_path = Path(download_path) / "antes_csv_LOTE.png"
    if screenshot_path.exists():
        try:
            from openpyxl.drawing.image import Image
            ws_screenshot = wb.create_sheet("Screenshot")
            img = Image(str(screenshot_path))
            ws_screenshot.add_image(img, 'A1')
        except Exception as e:
            print(f"Erro ao adicionar screenshot ao Excel: {e}")
    
    # Salvar Excel
    timestamp = time.strftime('%Y%m%d_%H%M%S')
    excel_path = Path(download_path) / f"relatorio_simexpress_{timestamp}.xlsx"
    wb.save(excel_path)
    print(f"Relatório Excel gerado: {excel_path}")


def main():
    parser = argparse.ArgumentParser(description="Bot para automação Simexpress")
    parser.add_argument('--pedidos', type=str, help='Caminho para arquivo Excel/CSV com pedidos (padrão: pedidos.xlsx)')
    args = parser.parse_args()

    Path(DOWNLOAD_PATH).mkdir(parents=True, exist_ok=True)

    logs = []  # Lista para coletar logs

    def log(msg):
        print(msg)
        logs.append(f"{time.strftime('%Y-%m-%d %H:%M:%S')} {msg}")
        with open(Path(DOWNLOAD_PATH) / "simexpress_log.txt", "a", encoding="utf-8") as f:
            f.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} {msg}\n")

    # Escolhe pedidos em ordem: CLI > .env (PEDIDOS_LOTE) > EXCEL_PATH
    if args.pedidos:
        pedidos_path = args.pedidos
        if os.path.exists(pedidos_path):
            log(f"Pedidos carregados do arquivo: {pedidos_path}")
            pedidos_text = _carregar_pedidos_do_arquivo(pedidos_path)
        else:
            log(f"Arquivo de pedidos não encontrado: {pedidos_path}")
            log("Pedidos carregados do .env (PEDIDOS_LOTE).")
            pedidos_text = _pedidos_do_env()
    elif os.getenv("PEDIDOS_LOTE"):
        log("Pedidos carregados do .env (PEDIDOS_LOTE).")
        pedidos_text = _pedidos_do_env()
    elif EXCEL_PATH and os.path.exists(EXCEL_PATH):
        log(f"Pedidos carregados do arquivo padrão: {EXCEL_PATH}")
        pedidos_text = _carregar_pedidos_do_arquivo(EXCEL_PATH)
    else:
        log("Nenhum arquivo de pedidos encontrado; usando .env (PEDIDOS_LOTE).")
        pedidos_text = _pedidos_do_env()

    # Valor final para envio ao portal
    global PEDIDOS_LOTE
    PEDIDOS_LOTE = pedidos_text

    # Verificar se há múltiplos pedidos
    pedidos_lista = [p.strip() for p in PEDIDOS_LOTE.split('\n') if p.strip()]
    
    # Sempre processar todos os pedidos de uma vez
    log(f"Processando {len(pedidos_lista)} pedido(s): {', '.join(pedidos_lista)}")
    try:
        resultado = processar_pedido_unico(PEDIDOS_LOTE.strip(), DOWNLOAD_PATH, log)
        if resultado:
            log("Processamento concluído com sucesso.")
        else:
            log("Processamento concluído com avisos.")
    except Exception as e:
        log(f"Erro durante processamento: {e}")

    # Gerar relatório em Excel
    gerar_relatorio_excel(logs, DOWNLOAD_PATH, pedidos_lista)

def _esperar_novo_csv(download_path, before_files, timeout=60):
    for _ in range(timeout):
        now_files = set(Path(download_path).glob('*.csv'))
        new_files = now_files - before_files
        if new_files:
            return max(new_files, key=lambda p: p.stat().st_mtime)
        time.sleep(1)
    return None


def processar_pedido_unico(pedidos_text, download_path, log_func):
    pedidos_lista = [p.strip() for p in pedidos_text.split('\n') if p.strip()]
    pedido_ident = "LOTE" if len(pedidos_lista) > 1 else pedidos_lista[0] if pedidos_lista else "UNKNOWN"
    # Criar novo driver para cada pedido (evita cache)
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--ignore-ssl-errors")
    options.add_argument("--ignore-certificate-errors-spki-list")
    options.add_argument("--ignore-ssl-errors-ignore-untrusted")
    prefs = {
        "download.default_directory": download_path,
        "download.prompt_for_download": False,
        "profile.default_content_setting_values.automatic_downloads": 1,
    }
    options.add_experimental_option("prefs", prefs)

    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        log_func(f"ChromeDriver path: {service.path}")
        log_func(f"ChromeDriver version: {webdriver.__version__}")
    except Exception as e:
        msg = ("Falha ao iniciar o ChromeDriver. Verifique se o Chrome está instalado, "
               "se a versão do Chrome é compatível com o ChromeDriver e se há permissão de execução. "
               f"Erro: {e}")
        log_func(msg)
        raise RuntimeError(msg)

    wait = WebDriverWait(driver, 40)

    def find_first(xpath_list):
        for xp in xpath_list:
            try:
                el = WebDriverWait(driver, 6).until(EC.element_to_be_clickable((By.XPATH, xp)))
                return el, xp
            except Exception:
                continue
        return None, None

    try:
        driver.get(URL)
        log_func("[1/8] Página inicial aberta")

        try:
            btn_acesso = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Acesso ao Sistema') or contains(., 'Acesso ao Sistema') or contains(@href, 'login') or contains(@class,'acesso')]")))
            btn_acesso.click()
            log_func("[2/8] Clique em Acesso ao Sistema")
        except Exception as e:
            log_func("[2/8] Acesso ao Sistema não clicado (pode já estar login): " + str(e))

        if XPATH_USUARIO:
            xpath_usuario = XPATH_USUARIO
        else:
            campo_usuario_el, xpath_usuario_found = find_first([
                "//input[@name='login' or @id='login' or @name='username' or @id='username' or @name='user' or @id='user']",
                "//input[contains(@placeholder,'Usuário') or contains(@placeholder,'user') or contains(@aria-label,'Usuário') or contains(@aria-label,'user')]",
                "//input[@type='text' or @type='email']",
            ])
            if not campo_usuario_el:
                raise RuntimeError("Campo de usuário não encontrado")
            xpath_usuario = xpath_usuario_found

        driver.find_element(By.XPATH, xpath_usuario).clear()
        driver.find_element(By.XPATH, xpath_usuario).send_keys(USUARIO)
        log_func("[3/8] Usuário inserido")

        if XPATH_SENHA:
            xpath_senha = XPATH_SENHA
        else:
            campo_senha_el, xpath_senha_found = find_first([
                "//input[@type='password' or @name='senha' or @id='senha']",
                "//input[contains(@placeholder,'Senha') or contains(@aria-label,'Senha') or contains(@class,'password')]",
            ])
            if not campo_senha_el:
                raise RuntimeError("Campo de senha não encontrado")
            xpath_senha = xpath_senha_found

        driver.find_element(By.XPATH, xpath_senha).clear()
        driver.find_element(By.XPATH, xpath_senha).send_keys(SENHA)
        log_func("[4/8] Senha inserida")

        if XPATH_ENTRAR:
            xpath_entrar = XPATH_ENTRAR
        else:
            botao_entrar_el, xpath_entrar_found = find_first([
                "//button[contains(translate(., 'ENTRAR', 'entrar'), 'entrar') or contains(.,'Entrar') or contains(.,'Login') or @type='submit']",
                "//input[@type='submit' and (contains(@value,'Entrar') or contains(@value,'Login'))]",
                "//button[.//svg or .//i or .//path][contains(@id,'login') or contains(@class,'login') or contains(@aria-label,'Entrar') or contains(@title,'Entrar') or contains(@name,'entrar') or contains(@name,'login') or @type='button']",
                "//form//button[.//svg or .//i or .//path or contains(translate(., 'ENTRAR', 'entrar'), 'entrar')]",
                "//form//a[.//svg or .//i or .//path or contains(translate(., 'ENTRAR', 'entrar'), 'entrar')]",
                "//div[contains(@class, 'login') or contains(@class, 'enter')]//button[1]",
            ])
            if not botao_entrar_el:
                try:
                    el = driver.execute_script("""
                        var texts = ['entrar','login','acessar'];
                        var nodes = Array.from(document.querySelectorAll('button,a,input,div,span'));
                        for (var i=0;i<nodes.length;i++){
                            var n = nodes[i];
                            var txt = (n.innerText || n.value || '').toString().toLowerCase().trim();
                            if(!txt) continue;
                            for(var j=0;j<texts.length;j++){
                                if(txt.indexOf(texts[j]) !== -1){
                                    n.click();
                                    return true;
                                }
                            }
                        }
                        return false;
                    """)
                    if not el:
                        raise RuntimeError("Botão Entrar não encontrado")
                except Exception:
                    raise RuntimeError("Botão Entrar não encontrado")
            else:
                xpath_entrar = xpath_entrar_found

        if not XPATH_ENTRAR:
            driver.find_element(By.XPATH, xpath_entrar).click()
        log_func("[5/8] Login enviado")

        wait.until(EC.url_changes(URL))
        log_func("[6/8] Login confirmado - URL mudou")
        time.sleep(2)

        # agora clique em Pedidos na tela pós-login
        menu_pedidos = wait.until(EC.element_to_be_clickable((By.XPATH,
            "//a[contains(normalize-space(.),'Pedidos') or contains(translate(normalize-space(.),'PEDIDOS','pedidos'),'pedidos') or contains(@href,'pedidos') or contains(@class,'pedido')]"
        )))
        menu_pedidos.click()
        log_func("[7/8] Clique em Pedidos")
        time.sleep(2)

        # o clique no menu pai pode apenas abrir submenu; aguardar o link 'Em Lote' aparecer
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, "//a[contains(normalize-space(.),'Em Lote') or contains(@href,'pedidos.php') or contains(translate(normalize-space(.),'EM LOTE','em lote'),'em lote')]")))
        except Exception:
            time.sleep(1)

        sub_em_lote = wait.until(EC.element_to_be_clickable((By.XPATH,
            "//a[contains(normalize-space(.),'Em Lote') or contains(translate(normalize-space(.),'EM LOTE','em lote'),'em lote') or contains(@href,'lote') or contains(@class,'lote')]"
        )))
        sub_em_lote.click()
        log_func("[8/8] Clique em Em Lote")
        time.sleep(2)

        # higiene de sessão local do site (limpar state de componentes, mas manter sessão)
        driver.execute_script("window.localStorage.clear(); window.sessionStorage.clear();")
        # NOTE: não limpar cookies após login, pois kill session e volta para login

        textarea_pedidos = wait.until(EC.visibility_of_element_located((By.XPATH, "//textarea[@id='pedidoLote' or contains(@placeholder,'cada item') or @name='pedidos'] | //textarea")))
        textarea_pedidos.clear()
        for pedido in pedidos_lista:
            textarea_pedidos.send_keys(pedido)
            textarea_pedidos.send_keys(Keys.RETURN)
        log_func(f"Textarea preenchido com: {pedidos_text}")
        time.sleep(3)

        # Clicar no botão Consultar e aguardar a página aplicar filtros
        botao_consultar = wait.until(EC.element_to_be_clickable((By.ID, "btnFiltrar")))
        driver.execute_script("arguments[0].click();", botao_consultar)
        log_func("Botão Consultar clicado via JavaScript")
        time.sleep(10)  # Aguardar os resultados carregarem

        # Limpar CSVs antigos antes de buscar o novo
        existing_csvs = set(Path(download_path).glob('*.csv'))


        # Screenshot antes de clicar no botão
        driver.save_screenshot(str(Path(download_path) / f"antes_csv_{pedido_ident}.png"))

        # Salvar page_source antes de clicar
        with open(Path(download_path) / f"antes_csv_{pedido_ident}.html", 'w', encoding='utf-8') as f:
            f.write(driver.page_source)

        botao_csv = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(translate(., 'CSV', 'csv'), 'csv') or contains(., 'Download') or contains(., 'Excel')] | //a[contains(translate(., 'CSV', 'csv'), 'csv') or contains(., 'Download')]")))

        # Remover CSV anteriores para evitar confusão de arquivo antigo
        for oldf in existing_csvs:
            try:
                oldf.unlink()
            except Exception:
                pass

        botao_csv.click()

        # aguardar o arquivo CSV aparecer na pasta de downloads e copiar para workspace/downloads.csv
        try:
            csv_found = _esperar_novo_csv(download_path, existing_csvs, timeout=60)
            if csv_found is None:
                csv_files = list(Path(download_path).glob('*.csv'))
                if csv_files:
                    csv_found = max(csv_files, key=lambda p: p.stat().st_mtime)

            dest_dir = Path(download_path)
            dest_dir.mkdir(parents=True, exist_ok=True)

            if csv_found:
                dest_path = dest_dir / f"SIMEXPRESS_{pedido_ident}.csv"
                if str(csv_found) != str(dest_path):
                    shutil.copy(str(csv_found), str(dest_path))
                else:
                    dest_path = csv_found
                log_func(f"CSV copiado para: {dest_path}")

                # Validação: verificar se os pedidos aparecem no CSV
                try:
                    df_resultado = pd.read_csv(str(dest_path))
                    if 'Pedido Cliente' in df_resultado.columns:
                        pedidos_encontrados = set(df_resultado['Pedido Cliente'].dropna().astype(str).str.strip('"').str.strip())
                        pedidos_faltando = [p for p in pedidos_lista if p not in pedidos_encontrados]
                        if not pedidos_faltando:
                            log_func(f"Validação OK: Todos os pedidos {pedidos_lista} encontrados no CSV")
                            return True
                        else:
                            log_func(f"Aviso: Pedidos não encontrados no CSV: {pedidos_faltando}")
                            # Imprimir resumo do CSV mesmo com avisos
                            try:
                                df = pd.read_csv(str(dest_path))
                                log_func(f"CSV contém {len(df)} linhas de dados.")
                                log_func(f"Colunas: {', '.join(df.columns)}")
                            except Exception as e:
                                log_func(f"Erro ao ler CSV para resumo: {e}")
                            return False
                    else:
                        log_func("Aviso: Coluna 'Pedido Cliente' não encontrada no CSV para validação")
                        return False
                except Exception as e:
                    log_func(f"Erro na validação do CSV: {e}")
                    return False
            else:
                log_func("Aviso: nenhum arquivo CSV encontrado no diretório de downloads dentro do tempo esperado")
                return False
        except Exception as e:
            log_func(f"Erro ao copiar CSV para pasta workspace: {e}")
            return False

        log_func("[9/9] CSV acionado e processo concluído")
        # Imprimir resumo do CSV
        try:
            df = pd.read_csv(str(dest_path))
            log_func(f"CSV contém {len(df)} linhas de dados.")
            log_func(f"Colunas: {', '.join(df.columns)}")
        except Exception as e:
            log_func(f"Erro ao ler CSV para resumo: {e}")
        time.sleep(30)  # Manter navegador aberto por 30 segundos para visualização

    except Exception as ex:
        driver.save_screenshot(str(Path(download_path) / f"simexpress_erro_{pedido_ident}.png"))
        try:
            with open(Path(download_path) / f"simexpress_erro_{pedido_ident}_page.html", 'w', encoding='utf-8') as hf:
                hf.write(driver.page_source)
            log_func(f"Salvo page_source em: {Path(download_path) / f'simexpress_erro_{pedido_ident}_page.html'}")
        except Exception:
            log_func("Falha ao salvar page_source de erro")
        log_func(f"Erro durante automação dos pedidos {pedido_ident}: {ex}")
        raise

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
